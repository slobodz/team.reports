from openpyxl import load_workbook
from openpyxl.drawing.image import Image as exImage
from openpyxl.styles import Alignment, Border, PatternFill, Side, Font
import json
from PIL import Image
from io import BytesIO
import os
import shutil

class ExcelFile:
    """ Load the template xlsx file which contains product_code list
    and update appropriate cells using the data from the api call
    """

    NOT_FOUND_ITEM = '<not found>'
    PRODUCT_CODE = 'product_code'
    PHOTO_CODE = 'image'
    TEMPLATE_FILE_NAME = 'template.xlsx'
    TEMPLATE_SHEET_NAME = 'products'
    TEMP_IMG_FILE_NAME = 'temp.jpeg'
    TEMP_IMG_FOLDER_NAME = 'teamreports_temp_images_folder'

    def __init__(self, filename=TEMPLATE_FILE_NAME, sheet=TEMPLATE_SHEET_NAME, target_filename=None):
        self.filename = filename
        self.target_filename = target_filename if target_filename is not None else filename
        self.wb = load_workbook(self.filename)
        self.sheet = self.wb[sheet]
        

        self.column_names_dict = {column_name.value: column_name.column
                                for column_name in self.sheet[1]} #for now assume column names are in 1st row

        if self.PRODUCT_CODE not in self.column_names_dict:
            raise KeyError("Column names not in first row")

        self.product_codes_dict = {
            product_code.value: product_code.row
            for product_code in self.sheet[self.column_names_dict[self.PRODUCT_CODE]]
            if product_code.row != 1 and product_code.value is not None
            }

        if not os.path.isdir(self.TEMP_IMG_FOLDER_NAME):
            os.mkdir(self.TEMP_IMG_FOLDER_NAME)

    def update_line(self, product, row=None):
        _current_row = self.product_codes_dict[product[self.PRODUCT_CODE]] if row is None else row

        self.sheet.row_dimensions[_current_row].height = 150 #experimental  

        for item in self.column_names_dict:
            _current_column = self.column_names_dict[item]
            _current_cell = _current_column + str(_current_row)

            #dimensions
            #self.sheet.column_dimensions[_current_column].width = 30 #experimental

            #alignment
            # self.sheet[_current_cell].alignment = Alignment(horizontal='left', vertical='center')

            #fill
            # if item in ['available', 'default_price', 'price']:
            #     self.sheet[_current_cell].fill = PatternFill(patternType='solid', start_color='D9D9D9')

            #border
            # self.sheet[_current_cell].border = Border(
            #                                             left=Side(border_style='thin', color='000000'),
            #                                             right=Side(border_style='thin', color='000000'),
            #                                             top=Side(border_style='thin', color='000000'),
            #                                             bottom=Side(border_style='thin', color='000000')
            #                                         )

            #bold
            # if item in ['available', 'default_price', 'price']:
            #     self.sheet[_current_cell].font = Font(bold=True)            

            if(item == self.PHOTO_CODE and item in product):        

                img = Image.open(BytesIO(product[item])) #change bytes from request into image obj
                img_path = self.TEMP_IMG_FOLDER_NAME + '/' + product['file_name']
                img.save(img_path) #save image on the disk
                img = exImage(img_path) #locate image using openpyxl
                self.sheet.column_dimensions[_current_column].width = (img.width/7) #experimental
                # self.sheet.row_dimensions[_current_row].height = (img.height *0.75) #experimental  
                self.sheet.add_image(img, _current_cell) #insert image

                     
            else:
                self.sheet[_current_cell] = product[item] if item in product else self.NOT_FOUND_ITEM


    def update_file_with_selected_products(self, product_list):
        for product in product_list:
            self.update_line(product)


    def list_all_products(self):
        return [product_code for product_code in self.product_codes_dict]


    def save_file(self, skip_delete=False):
        self.wb.close()
        self.wb.save(self.target_filename)
        if not skip_delete:
            if os.path.isdir(self.TEMP_IMG_FOLDER_NAME):
                shutil.rmtree(self.TEMP_IMG_FOLDER_NAME)


    def update_all_products(self, product_list):
        _product_code_column = self.column_names_dict[self.PRODUCT_CODE]

        for row, product in enumerate(product_list, 2):
            self.sheet[_product_code_column + str(row)] = product[self.PRODUCT_CODE]
            self.update_line(product, row=row)

